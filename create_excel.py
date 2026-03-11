from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: Master Rifugio List ──
ws = wb.active
ws.title = "Master Rifugio List"

headers = [
    "#", "Rifugio", "Trail Position", "Elevation", "Region / Route", "Vibe",
    "Rating (1-5)", "Price (Half-Board pp/night)", "Private Rooms?",
    "Booking Lead Time", "Website URL"
]

# Ordered NORTH → SOUTH along AV1 corridor (+ nearby detours slotted where you'd access them)
# Includes city hotels for pre/post trek
data = [
    # ── PRE-TRIP: Venice ──
    [1, "Ca' Del Sole Venezia", "Pre-trek — Venice (night before travel to Dolomites)", "5 m", "Venice / San Marco",
     "Venice base before the trek. Central San Marco location. Good launching point — train to Dobbiaco next morning.",
     3, "~€100-160/night", "Yes",
     "Book 1-2 months out. Standard city hotel availability.",
     "https://ca-del-sol-venezia.venetohotelsweb.com/en/"],

    # ── PRE-TRIP: Dobbiaco ──
    [2, "Hotel Toblacherhof", "Pre-trek — Dobbiaco (night before AV1 start)", "1,243 m", "Dobbiaco (Toblach)",
     "3-star family hotel. Spa, sauna, great breakfast. 20 min from Lago di Braies. 8.4/10 on Booking.com. Perfect night-before base.",
     4, "~€158+/night", "Yes (standard + superior rooms)",
     "Book 2-4 months out. Popular summer base.",
     "https://www.toblacherhof.com/en/"],

    # ── FAR NORTH: Tre Cime area (optional pre-AV1 start or side trip) ──
    [3, "Rifugio Auronzo", "Tre Cime trailhead (side trip or pre-AV1)", "2,320 m", "Tre Cime di Lavaredo",
     "The gateway. Right at the Tre Cime trailhead. More accessible (road access), busier during the day but empties out at night. Great launching pad for sunrise at Tre Cime.",
     4, "€90-100 dorm / €220-240 double", "Yes (doubles, triples)",
     "Book 4-6 months out. Peak Jul-Aug books fastest.",
     "https://rifugioauronzo.it/en/booking-room/"],
    [4, "Rifugio Locatelli (Drei Zinnen)", "Tre Cime circuit (side trip or pre-AV1)", "2,405 m", "Tre Cime di Lavaredo",
     "The postcard. Sits directly beneath the three most famous peaks in the Dolomites. The view from the terrace is unreal. Gets ~10,000 booking requests/year. Cash only.",
     5, "~€85-95 half-board (dorm)", "Limited (4-bed rooms)",
     "Book the moment reservations open (Feb-Mar). Expect 3-week response time. 6+ months out minimum.",
     "https://www.dreizinnenhuette.com/rifugio-locatelli.php"],

    # ── NORTH: Braies / Fanes-Sennes-Braies (AV1 Stages 1-2) ──
    [5, "Hotel Prato Piazza (Plätzwiese)", "Near Braies — pre-AV1 or Night 0", "2,000 m", "Braies area",
     "The meadow dream. Wide alpine meadow with panoramic views. More hotel than hut. Great for families or those wanting comfort at the start/end of a trek.",
     4, "~€90-120 half-board (est.)", "Yes (hotel rooms)",
     "Book 4-6 months out. Popular with day-trippers too.",
     "https://www.pratopiazza.com/"],
    [6, "Rifugio Biella", "AV1 Stage 1 end", "2,327 m", "Sennes Plateau / AV1",
     "The minimalist. Basic but well-positioned as the first AV1 overnight from Lago di Braies. Dorms only. Functional stop for through-hikers.",
     2, "~€55-65 half-board (est.)", "No (dorms only)",
     "Book 2-4 months out.",
     "Contact via CAI"],
    [7, "Rifugio Sennes", "AV1 Stage 1-2 area", "2,126 m", "Fanes-Sennes-Braies Park / AV1",
     "The plateau gem. Historic hut on a wide-open alpine plateau. Hearty South Tyrolean food, homemade strudel, local wine. Peaceful, wide-sky feeling.",
     4, "~€70-85 half-board (est.)", "Yes (limited)",
     "Book 4-6 months out via website contact form.",
     "https://www.sennes.com/en/"],
    [8, "Rifugio Fodara Vedla", "AV1 Stage 1-2 area (Sennes plateau)", "1,980 m", "Fanes-Sennes-Braies Park / AV1",
     "Family-run gem on the Sennes plateau. 12 rooms, 42 beds. Garden-fresh salads, handmade pasta. Traditional South Tyrolean warmth. Mutschlechner family.",
     4, "€83+ half-board", "Yes (12 rooms)",
     "Book 4-6 months out. Open mid-Jun to early Nov.",
     "https://www.fodara.it/en"],
    [9, "Rifugio Pederü", "AV1 valley base (Stage 2 start)", "1,548 m", "Fanes-Sennes-Braies Park / AV1",
     "The launch pad. Family-run alpine guesthouse in classic South Tyrolean wood style. Lower elevation = warmer, more comfortable. Perfect Day 1 or Day 0 base.",
     3, "~€70-85 half-board (est.)", "Yes (traditional wood rooms)",
     "Book 3-5 months out. 3-night min may apply.",
     "https://www.pederue.it/en"],
    [10, "Rifugio Lavarella (Berghütte)", "AV1 Stage 2 option", "2,042 m", "Fanes-Sennes-Braies Park / AV1",
     "The cozy upgrade. Family-run, South Tyrolean warmth. Has a sauna AND a microbrewery. Comfortable rooms. Great base for exploring the Fanes plateau.",
     4, "~€80-100 half-board (est.)", "Yes (doubles + dorms)",
     "Book 4-6 months out. Online booking at lavarella.it.",
     "https://lavarella.it/en/book-online"],
    [11, "Rifugio Fanes", "AV1 Stage 2 end", "2,060 m", "Fanes-Sennes-Braies Park / AV1",
     "The all-rounder. Beautiful alpine plateau setting. Range of room types from dorms to premium doubles with private bath. Good food, accepts cards. Feels more like a mountain hotel than a hut.",
     4, "€80-95 dorm / €103-135 private", "Yes (doubles, quads, premium w/ private bath)",
     "Book 4-6 months out. Online booking available.",
     "https://www.rifugiofanes.com/en/rates-request/prices-conditions"],

    # ── MID-NORTH: Lagazuoi / Scotoni area (AV1 Stage 3) ──
    [12, "Rifugio Scotoni", "AV1 Stage 3 — en route to Lagazuoi", "1,985 m", "Alta Badia / near AV1",
     "The foodie stop. Over 50 years old. Famous for indoor barbecue and a wine list with 400+ labels. Hidden valley location in Lagazuoi area. Eat like a king at altitude.",
     4, "~€75-90 half-board (est.)", "Yes (limited)",
     "Book 3-5 months out. Contact directly.",
     "https://www.scotoni.it/en/how-to-reach-us/from-fanes-dolomites"],
    [13, "Rifugio Lagazuoi", "AV1 Stage 3 end (highest point)", "2,752 m", "Falzarego Pass / AV1",
     "The crown jewel. Highest rifugio in the Dolomites. 360° panorama of Marmolada, Tofane, Cinque Torri. Finnish sauna at altitude. Sunrises are legendary. Cable car access = easy bail option.",
     5, "€80 dorm / €125 double (B&B) + €35 dinner", "Yes (double, twin w/ balcony, triple, quad)",
     "Book 6-9 months out; opens Sept/Oct for next summer. Sells out fast.",
     "https://rifugiolagazuoi.com/index_en.php"],
    [14, "Rifugio Col Gallina", "AV1 Stage 3-4 — Falzarego Pass", "2,055 m", "Falzarego Pass / AV1",
     "At Falzarego Pass, center of Lagazuoi/5 Torri/Giau area. Double rooms w/ private bath, WiFi, towels. 50% deposit. Solid alternative to Lagazuoi.",
     3, "~€75-90 half-board (est.)", "Yes (doubles + dorms w/ private bath)",
     "Book 3-5 months out. Contact directly.",
     "https://www.rifugiocolgallina.com/en/"],

    # ── MID: Cinque Torri / Averau / Nuvolau area (AV1 Stage 4) ──
    [15, "Rifugio Giussani", "Detour — Tofane (accessible from Falzarego)", "2,580 m", "Tofane Group / near AV1",
     "The high camp. Dramatic position beneath the Tofane peaks. Good base for via ferrata enthusiasts. Slightly off the beaten path of AV1 proper.",
     3, "€73 (€58.50 Alpine Club members)", "Limited",
     "Book 3-5 months out.",
     "https://www.rifugiogiussani.com/en/accommodation-rates/"],
    [16, "Rifugio Cinque Torri (5 Torri)", "AV1 Stage 4 — near Averau", "2,137 m", "Cinque Torri",
     "The rock garden. Surrounded by the otherworldly Cinque Torri rock towers. Great for climbers. Shuttle access in summer (€10).",
     3, "~€70-80 half-board (est.)", "Limited",
     "Book 3-5 months out.",
     "https://rifugio5torri.it/en/5-torri.html"],
    [17, "Rifugio Averau", "AV1 Stage 4 option", "2,413 m", "Cinque Torri / AV1",
     "The balcony. Perched with views of Cinque Torri rock formations and surrounding peaks. Good food, nice terrace. Solid mid-route stop with character.",
     4, "~€75-90 half-board (est.)", "Yes (limited doubles)",
     "Book 4-6 months out.",
     "https://www.rifugioaverau.it/"],
    [18, "Rifugio Nuvolau", "AV1 Stage 4 end", "2,575 m", "Monte Nuvolau / AV1",
     "The OG. Oldest rifugio in the Dolomites (est. 1883). Tiny, perched on a summit, feels like floating above the clouds. No showers, no private rooms — pure mountain soul. Sunset spot of a lifetime.",
     5, "~€70-80 half-board (est.)", "No (dorms only)",
     "Bookings open Feb 1 for that summer. Open Jun 20-Sep 20. Book immediately when window opens.",
     "https://rifugionuvolau.it/en/"],

    # ── MID: Cortina side / Croda da Lago / Giau (detour or AV1 variant) ──
    [19, "Rifugio Croda da Lago (Palmieri)", "Detour — Lago Federa (from Cortina side)", "2,046 m", "Lake Federa / near Cortina",
     "The hidden lake. Sits on the shore of stunning turquoise Lago Federa. Barrel sauna. Great food with buffet breakfast. Surrounded by larches that turn gold in September. Feels like a painting.",
     5, "€73 (€58.50 CAI/DAV members)", "No (6/9/15-bed dorms)",
     "Book 3-6 months out. Online booking with real-time availability.",
     "https://www.crodadalago.it/en/prices-and-stay/"],
    [20, "Rifugio Fedare", "AV1 variant — Passo Giau", "2,000 m", "Passo Giau",
     "The pass stop. At scenic Giau Pass. Road accessible. Good views but more of a waypoint than a destination.",
     3, "€80", "Yes",
     "Book 3-5 months out.",
     "https://valfiorentinadolomiti.it/en/Rifugio-Fedare-Passo-Giau/The_rooms/Room-prices.html"],

    # ── MID-SOUTH: Pelmo area (AV1 Stage 5) ──
    [21, "Rifugio Città di Fiume", "AV1 Stage 5 — foot of Pelmo", "1,918 m", "Monte Pelmo / AV1",
     "The Pelmo basecamp. At the foot of the massive Pelmo monolith. Meadow setting with outdoor dining. Small (20 beds), intimate. Good food.",
     3, "~€65-75 half-board (est.)", "No (4 shared rooms)",
     "Book 3-5 months out. Open mid-Jun to late Sep.",
     "https://rifugiocittadifiume.it/?lang=en"],
    [22, "Rifugio Alpino Aquileia", "Detour — near Pelmo (slight AV1 detour)", "1,583 m", "Selva di Cadore / near AV1",
     "Family-run, 25 beds. Homemade pasta, dumplings, game dishes. Slight detour from AV1 but reviewers say it's worth it. Knowledgeable owner, great route advice.",
     3, "~€65-75 half-board (est.)", "Yes (limited)",
     "Book 3-5 months out.",
     "https://www.rifugioaquileia.it/"],
    [23, "Rifugio Passo Staulanza", "AV1 Stage 5-6 transition", "1,783 m", "Staulanza Pass / AV1",
     "The comfort stop. Road-accessible, 13 rooms with en-suite bathrooms. More hotel than hut. Restaurant with traditional cuisine. Good reset point for a real shower and proper bed.",
     3, "~€70-85 half-board (est.)", "Yes (13 rooms, all en-suite)",
     "Book 2-4 months out. Easier availability due to road access.",
     "https://www.staulanza.it/prenotazioni.php?lang=en"],

    # ── SOUTH: Civetta Group (AV1 Stages 6-8) ──
    [24, "Rifugio Coldai (Sonino al Coldai)", "AV1 Stage 6 — Civetta north", "2,135 m", "Civetta Group / AV1",
     "The wild card. 93 beds so availability is better. Sits near Lago Coldai with Civetta views. Less famous than Tissi but arguably just as beautiful. Great for a quieter experience.",
     4, "~€65-75 half-board (est.)", "Limited",
     "Book 3-5 months out. Online booking available.",
     "https://rifugiocoldai.com/index_en.php"],
    [25, "Rifugio Tissi", "AV1 Stage 7 — Civetta south", "2,250 m", "Monte Civetta / AV1",
     "The sunset king. Civetta's 1,000m vertical wall turns blood-red at sunset — one of the most dramatic views on any AV route. Remote feel, small, intimate.",
     5, "~€65-75 half-board (est.)", "No (dorms)",
     "Book 4-6 months out. Less known = slightly easier than Lagazuoi/Locatelli.",
     "Contact via CAI section"],
    [26, "Rifugio Vazzoler", "AV1 Stage 7-8 — below Civetta", "1,714 m", "Civetta Group / AV1",
     "The lowkey legend. Beautifully situated beneath Civetta. Cash only, €30 deposit. Old-school rifugio vibe. Less polished, more authentic.",
     3, "~€60-70 half-board (est.)", "Limited",
     "Book 3-5 months out. Cash only, non-refundable deposit.",
     "https://rifugiovazzoler.com/Booking/index_en.php"],
    [27, "Rifugio Carestiato", "AV1 Stage 8 — toward Agordo", "2,760 m", "Marmolada area",
     "The high point. Near the Marmolada glacier area. Online booking available. Deposit required.",
     3, "~€65-75 half-board (est.)", "Limited",
     "Book 3-5 months out.",
     "https://www.rifugiocarestiato.com/Booking/index_en.php"],

    # ── FAR SOUTH: Bellunesi Dolomites (AV1 Stages 9-11) ──
    [28, "Rifugio Sommariva al Pramperet", "AV1 Stage 9 — Val Pramper", "1,857 m", "Val Pramper / AV1 South",
     "The time capsule. Built in 1923. 25 beds, 6-8 person rooms. Winter bivouac available. Deep in Val Pramper — feels like stepping back 100 years.",
     3, "~€55-65 half-board (est.)", "No (shared 6-8 bed rooms)",
     "Book 2-4 months out. Open Jun 20-Sep 20.",
     "https://www.rifugiosommarivaalpramperet.it/rifugio.asp"],
    [29, "Rifugio Pian de Fontana", "AV1 Stage 10 — deep Bellunesi", "1,632 m", "Dolomiti Bellunesi / AV1 South",
     "The chef's table. Known for excellent food (manager Gavino is a local legend). Remote, in the wild Bellunesi Dolomites. Far fewer tourists. True off-the-beaten-path.",
     4, "~€55-65 half-board (est.)", "No (dorms)",
     "Book 2-4 months out. Remote = less demand.",
     "Phone: +39 0437 1956135"],

    # ── POST-TREK: Belluno / Brenta area ──
    [30, "Tenuta Contarini", "Post-trek — Brenta Valley (wind-down before Venice)", "120 m", "Campolongo sul Brenta",
     "17-room country hotel in Brenta Valley, garden setting across 3 houses. Good post-trek wind-down spot before heading back to Venice.",
     3, "~€80-120/night", "Yes (doubles, triples, family rooms)",
     "Book 1-2 months out. Standard hotel availability.",
     "https://www.contarini.it/en"],

    # ── OFF-ROUTE: Rosengarten (separate massif, not on AV1 but worth noting) ──
    [31, "Rifugio Re Alberto (Garlhutte)", "Off-AV1 — Rosengarten massif (separate trip/add-on)", "2,621 m", "Rosengarten / Catinaccio",
     "The adventure pick. Accessed via dramatic steel ladders (light via ferrata). Secluded valley, far from crowds. Book via WhatsApp/email only. Earn your dinner.",
     4, "~€60-70 half-board (est.)", "No (dorms)",
     "Book 2-4 months out. Less demand = more flexibility. Cash only.",
     "Contact via email/WhatsApp"],
]

# ── Styles ──
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2D5F2D", end_color="2D5F2D", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

star5_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # gold tint
star4_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # light green
star3_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # light gray
star2_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # gray

data_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

name_font = Font(name="Calibri", bold=True, size=11)
link_font = Font(name="Calibri", color="1155CC", underline="single", size=10)
normal_font = Font(name="Calibri", size=10)
rating_font = Font(name="Calibri", size=14)

# ── Write headers ──
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# ── Write data ──
star_map = {5: "★★★★★", 4: "★★★★☆", 3: "★★★☆☆", 2: "★★☆☆☆", 1: "★☆☆☆☆"}
fill_map = {5: star5_fill, 4: star4_fill, 3: star3_fill, 2: star2_fill, 1: None}

for row_idx, row_data in enumerate(data, 2):
    rating = row_data[6]  # Rating is now index 6 (shifted by Trail Position column)
    row_fill = fill_map.get(rating)

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border

        if col_idx == 7:  # Rating column — use stars
            cell.value = star_map.get(value, str(value))
            cell.font = rating_font
            cell.alignment = center_align
        elif col_idx == 2:  # Name column — bold
            cell.value = value
            cell.font = name_font
            cell.alignment = data_align
        elif col_idx == 11:  # URL column — hyperlink
            if value.startswith("http"):
                display = value.replace("https://", "").replace("http://", "").rstrip("/")
                if len(display) > 40:
                    display = display[:37] + "..."
                cell.value = display
                cell.hyperlink = value
                cell.font = link_font
            else:
                cell.value = value
                cell.font = normal_font
            cell.alignment = data_align
        elif col_idx in (1, 4):  # # and Elevation — center
            cell.value = value
            cell.font = normal_font
            cell.alignment = center_align
        elif col_idx == 3:  # Trail Position — italic accent
            cell.value = value
            cell.font = Font(name="Calibri", italic=True, size=10, color="555555")
            cell.alignment = data_align
        else:
            cell.value = value
            cell.font = normal_font
            cell.alignment = data_align

        if row_fill:
            cell.fill = row_fill

# ── Column widths ──
col_widths = {
    1: 4,    # #
    2: 32,   # Rifugio
    3: 38,   # Trail Position
    4: 12,   # Elevation
    5: 30,   # Region
    6: 65,   # Vibe
    7: 14,   # Rating
    8: 35,   # Price
    9: 30,   # Private Rooms
    10: 50,  # Booking Lead Time
    11: 42,  # URL
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:K{len(data) + 1}"

# ── Sheet 2: Booking Timeline ──
ws2 = wb.create_sheet("Booking Timeline")

timeline_headers = ["Booking Window", "What to Do"]
timeline_data = [
    ["Sept-Oct (year before)", "Reservations open for many rifugios. Start booking the big names (Lagazuoi, Locatelli, Nuvolau)."],
    ["Feb 1", "Nuvolau bookings open. Act immediately."],
    ["6-9 months out", "Book all 5-star and 4-star huts. AV1 sells out faster than Tour du Mont Blanc."],
    ["3-5 months out", "Book 3-star huts and fill gaps. Remote/southern huts still available."],
    ["1-2 months out", "Scraps only for popular huts. Remote Bellunesi huts may still have space."],
]

for col_idx, header in enumerate(timeline_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row_idx, row_data in enumerate(timeline_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Calibri", bold=(col_idx == 1), size=11)
        cell.alignment = data_align
        cell.border = thin_border

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 80
ws2.freeze_panes = "A2"

# ── Sheet 3: Notes ──
ws3 = wb.create_sheet("Notes & Tips")

notes = [
    "Prices marked (est.) are estimates based on comparable huts in the same tier. Contact directly for exact 2026 rates.",
    "Half-board = dinner + bed + breakfast (standard at nearly all rifugios).",
    "Cash only is common — bring euros. Some newer/renovated huts accept cards (Fanes, Croda da Lago, Lagazuoi).",
    "Tourist tax of ~€1.50/person/night is added on-site at most huts.",
    "CAI/DAV/Alpine Club members get 20-25% discounts at many huts.",
    "Showers cost €5-8 extra at most huts (some have no showers at all).",
    "No centralized booking system — each hut is booked individually via their own website, email, or phone.",
    "Best hiking season: July to September.",
    "Nearest airports: Venice (VCE) or Innsbruck (INN).",
]

cell = ws3.cell(row=1, column=1, value="Key Notes & Tips")
cell.font = Font(name="Calibri", bold=True, size=14, color="2D5F2D")

for i, note in enumerate(notes, 3):
    bullet = ws3.cell(row=i, column=1, value=f"• {note}")
    bullet.font = Font(name="Calibri", size=11)
    bullet.alignment = Alignment(wrap_text=True)

ws3.column_dimensions["A"].width = 100

# ── Save ──
output_path = r"c:\Users\DanPeterson\Dans Projects\Dolomites_Itinerary\Dolomites_Rifugio_Master_List_v2.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
